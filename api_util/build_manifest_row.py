#!/usr/bin/env python3
"""
build_manifest_row.py  –  Extract ordered text from one CSV/XLSX file,
write it to a temp text file, and print one TSV row to stdout:

    doc_id <TAB> page_count <TAB> /path/to/text_file
"""

import argparse
import csv
import os
import sys
from pathlib import Path

# Inject project root into sys.path so the vendored `atrium_document` resolves when this
# script is invoked directly (api_1_manifest.sh runs `python3 api_util/build_manifest_row.py`,
# which puts api_util/ — not the repo root — on sys.path). Same idiom as
# api_util/summarize_nt_udp.py.
_project_root = Path(__file__).resolve().parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

from atrium_document import canonical_doc_id  # noqa: E402

csv.field_size_limit(sys.maxsize)

try:
    import openpyxl
except ImportError:
    openpyxl = None


def _read_csv(file_path):
    entries = []
    try:
        with open(file_path, "r", encoding="utf-8") as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                try:
                    p = int(row.get("page_num", 0) or 0)
                except (ValueError, TypeError):
                    p = 0
                try:
                    ln = int(row.get("line_num", 0) or 0)
                except (ValueError, TypeError):
                    ln = 0
                text = (row.get("text") or "").strip()
                if text:
                    entries.append({"p": p, "l": ln, "text": text})
    except Exception as exc:
        print(f"[Error] reading CSV {file_path}: {exc}", file=sys.stderr)
    return entries


def _read_xlsx(file_path):
    if openpyxl is None:
        print("[Error] openpyxl is required for .xlsx files.", file=sys.stderr)
        return []
    entries = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            if not headers or "text" not in headers:
                continue
            text_idx = headers.index("text")
            page_idx = headers.index("page_num") if "page_num" in headers else -1
            line_idx = headers.index("line_num") if "line_num" in headers else -1
            for row in ws.iter_rows(min_row=2, values_only=True):
                text_val = row[text_idx]
                text = str(text_val).strip() if text_val is not None else ""
                if not text:
                    continue
                try:
                    p = int(row[page_idx]) if page_idx != -1 and row[page_idx] is not None else 0
                except (ValueError, TypeError):
                    p = 0
                try:
                    ln = int(row[line_idx]) if line_idx != -1 and row[line_idx] is not None else 0
                except (ValueError, TypeError):
                    ln = 0
                entries.append({"p": p, "l": ln, "text": text})
    except Exception as exc:
        print(f"[Error] reading XLSX {file_path}: {exc}", file=sys.stderr)
    return entries


def get_sorted_text_and_page_count(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        entries = _read_csv(file_path)
    elif ext == ".xlsx":
        entries = _read_xlsx(file_path)
    else:
        return None, 0
    if not entries:
        return None, 0
    entries.sort(key=lambda x: (x["p"], x["l"]))
    page_count = max((e["p"] for e in entries), default=0)
    full_text = "\n".join(e["text"] for e in entries)
    return full_text, page_count


def main():
    default_text_dir = os.environ.get("TEMP_TXT_DIR", "./TEMP/TXT_EXTRACT")
    parser = argparse.ArgumentParser()
    parser.add_argument("input_file")
    parser.add_argument("--text-dir", default=default_text_dir)
    parser.add_argument("--doc-id", default=None, help="Logical document ID to use in manifest")
    args = parser.parse_args()

    if not os.path.isfile(args.input_file):
        print(f"[Error] File not found: {args.input_file}", file=sys.stderr)
        sys.exit(1)

    # canonical_doc_id(), not os.path.splitext(): this row IS the doc_id every later stage
    # inherits (UDP/<file>.conllu → NE/<file> → the document record), so a derivation that
    # strips only the last extension forks the whole run's identity on any multi-dot input
    # (issue atrium-project#10, D3).
    doc_id = args.doc_id or canonical_doc_id(args.input_file)
    full_text, page_count = get_sorted_text_and_page_count(args.input_file)

    if full_text is None:
        print(f"[Error] No text content found in {args.input_file}", file=sys.stderr)
        sys.exit(1)

    out_text_file = os.path.join(args.text_dir, f"{doc_id}.txt")
    os.makedirs(os.path.dirname(out_text_file), exist_ok=True)
    with open(out_text_file, "w", encoding="utf-8") as fh:
        fh.write(full_text)

    print(f"{doc_id}\t{page_count}\t{out_text_file}")


if __name__ == "__main__":
    main()
